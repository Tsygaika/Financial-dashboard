import openpyxl
import yadisk
import os

y = yadisk.YaDisk(token="TOKEN")
table_name = './Таблица финансов.xlsx'

def read_excel_with_merged_cells(filename, sheet_name):
    workbook = openpyxl.load_workbook(filename, data_only=True)
    sheet = workbook[sheet_name]

    merged_ranges = sheet.merged_cells.ranges

    merged_values = {}
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_cell = sheet.cell(row=min_row, column=min_col)

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = top_left_cell.value

    all_data = []
    for row in range(1, sheet.max_row + 1):
        row_data = []
        for col in range(1, sheet.max_column + 1):
            if (row, col) in merged_values:
                value = merged_values[(row, col)]
            else:
                cell = sheet.cell(row=row, column=col)
                value = cell.value
            row_data.append(value)
        all_data.append(row_data)

    return all_data


from subscripts.expenses import parse_expenses
rows = read_excel_with_merged_cells(table_name, 'Расходы')
parse_expenses(rows)

from subscripts.income import parse_income
rows = read_excel_with_merged_cells(table_name, 'Доходы')
parse_income(rows)

from subscripts.portfolio_refills import parse_refills
rows1 = read_excel_with_merged_cells(table_name, 'Счёт 1')
rows2 = read_excel_with_merged_cells(table_name, 'Счёт 2')
parse_refills(rows1, rows2)

from subscripts.deposits import parse_deposits
rows = read_excel_with_merged_cells(table_name, 'Вклады')
parse_deposits(rows)

from subscripts.stocks import parse_stocks
rows = read_excel_with_merged_cells(table_name, 'Акции')
parse_stocks(rows)

from subscripts.capital import parse_capital
rows = read_excel_with_merged_cells(table_name, 'Активы')
parse_capital(rows)


files = os.listdir('./output/')

for file in files:
    y.upload(f'./output/{file}', f'Databases/Datalens/{file}', overwrite = True)